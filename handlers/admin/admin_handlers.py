import os
from zipfile import ZipFile

import openpyxl
from aiogram import F
from aiogram import types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile
from aiogram.types import Message
from loguru import logger

from database.database import UserStart
from keyboards.admin_keyboards import admin_keyboard
from system.dispatcher import bot, ADMIN_USER_ID
from system.dispatcher import dp
from system.dispatcher import router


@dp.message(F.text == "/admin_start")
async def admin_start_handler(message: Message) -> None:
    user_id = message.from_user.id
    user_name = message.from_user.username
    user_first_name = message.from_user.first_name
    user_last_name = message.from_user.last_name
    user_date = message.date.strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"{user_id} {user_name} {user_first_name} {user_last_name} {user_date}")
    sign_up_text = (
        "👋 <b>Добро пожаловать, Администратор!</b>\n\n"
        "Рад видеть вас в панели управления ботом. Здесь вы можете управлять ботом, "
        "редактировать тексты и получать важные данные о пользователях.\n\n"
        "<b>Доступные функции:</b>\n"
        "🔧 <i>Редактирование меню бота и инструкций</i>\n"
        "📊 <i>Получение данных о пользователях, которые запустили бота</i>\n\n"
        "Чтобы начать, выберите нужную команду из меню ниже или введите её вручную.\n\n"
        "⬇️ Воспользуйтесь меню ниже для дальнейших действий."
    )
    await bot.send_message(message.from_user.id, sign_up_text, reply_markup=admin_keyboard(),
                           disable_web_page_preview=True, parse_mode="HTML")


@router.callback_query(F.data == "admin_commands")
async def admin_send_start(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработчик команды /start, он же пост приветствия 👋"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    """Админ панель"""
    if callback_query.from_user.id not in ADMIN_USER_ID:
        await bot.send_message(callback_query.from_user.id, text="❌ У вас нет прав на выполнение этой команды.")
        return

    admin_commands_text = (
        "👮‍♂️ <b>Административная панель</b>\n\n"
        "Добро пожаловать в панель управления. Здесь вы можете редактировать тексты бота "
        "и получать данные о пользователях.\n\n"

        "🔧 <b>Редактирование текстов:</b>\n"
        "✔️ /edit_main_menu - <i>Изменение текста главного меню бота.</i>\n"
        "✔️ /edit_instructions - <i>Изменение текста раздела «Инструкция».</i>\n"
        "✔️ /edit_check_out_the_warranty - <i>Изменение текста раздела «Проверить / оформить гарантию».</i>\n"
        "✔️ /edit_guarantee_chek - <i>Изменение текста раздела «Хочу заполнить гарантийный талон».</i>\n"
        "✔️ /edit_check_the_warranty_card - <i>Изменение текста раздела «Хочу проверить гарантийный талон».</i>\n\n"

        "📊 <b>Получение данных:</b>\n"
        "✔️ /get_users_who_launched_the_bot - <i>Выгрузка списка пользователей, которые запускали бота.</i>\n"
        "✔️ /get_warranty_cards - <i>Скачать сформированные гарантийные талоны.</i>\n"
        "✔️ /get_photos_goods - <i>Скачать фото товаров.</i>\n\n"

        "ℹ️ Чтобы вернуться в начальное меню, нажмите /admin_start."
    )

    await bot.send_message(callback_query.from_user.id, text=admin_commands_text, parse_mode="HTML")


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Дата запуска бота'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # Имя
        sheet.cell(row=index, column=4).value = order[3]  # Фамилия
        sheet.cell(row=index, column=5).value = order[4]  # Дата запуска бота

    return workbook  # Возвращаем объект workbook


# Функция для чтения данных из базы данных
def reading_from_database():
    # Извлекаем все записи из таблицы UserStart
    query = UserStart.select()

    # Формируем список из кортежей с данными для передачи в create_excel_file_start
    orders = [
        (user.telegram_id, user.telegram_username, user.user_first_name, user.user_last_name, user.user_date)
        for user in query
    ]

    return orders


def create_zip_archive_photos_goods():
    # Создаем ZIP-файл
    with ZipFile('archive.zip', 'w') as archive:
        # Перебираем все файлы в текущем каталоге
        for foldername, subfolders, filenames in os.walk('product_photo'):
            for filename in filenames:
                # Добавляем файл в архив
                archive.write(os.path.join(foldername, filename))
    file = 'archive.zip'
    return file


@router.message(Command("get_photos_goods"))
async def get_users_get_photos_goods(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    try:
        if message.from_user.id not in [535185511, 301634256, 244948554, 496261899]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        files = create_zip_archive_photos_goods()  # Создание ZIP-файла
        logger.info(f"Создан ZIP-файл: {files}")
        file = FSInputFile(files)
        text = (
            "📊 <b>Фото товаров пользователей</b>\n\n"
            "⬇️ <b>Скачайте файл</b>, чтобы просмотреть данные.\n\n"
            "Для возврата в начальное меню нажмите на /admin_start"
        )
        await bot.send_document(message.from_user.id, document=file, caption=text,
                                parse_mode="HTML")  # Отправка файла пользователю
        os.remove(files)  # Удаление файла
    except Exception as e:
        logger.error(f"Произошла ошибка при попытке отправить файл: {e}")
        await bot.send_message(message.from_user.id,
                               "⚠️ Произошла ошибка при обработке вашего запроса. Пожалуйста, попробуйте позже.")


def create_zip_archive():
    # Создаем ZIP-файл
    with ZipFile('archive.zip', 'w') as archive:
        # Перебираем все файлы в текущем каталоге
        for foldername, subfolders, filenames in os.walk('completed_form'):
            for filename in filenames:
                # Добавляем файл в архив
                archive.write(os.path.join(foldername, filename))
    file = 'archive.zip'
    return file


@router.message(Command("get_warranty_cards"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    try:
        if message.from_user.id not in [535185511, 301634256, 244948554, 496261899]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        files = create_zip_archive()  # Создание ZIP-файла
        logger.info(f"Создан ZIP-файл: {files}")
        file = FSInputFile(files)
        text = (
            "📊 <b>Заполненные гарантийные пользователей</b>\n\n"
            "⬇️ <b>Скачайте файл</b>, чтобы просмотреть данные.\n\n"
            "Для возврата в начальное меню нажмите на /admin_start"
        )
        await bot.send_document(message.from_user.id, document=file, caption=text,
                                parse_mode="HTML")  # Отправка файла пользователю
        os.remove(files)  # Удаление файла
    except Exception as e:
        logger.error(f"Произошла ошибка при попытке отправить файл: {e}")
        await bot.send_message(message.from_user.id,
                               "⚠️ Произошла ошибка при обработке вашего запроса. Пожалуйста, попробуйте позже.")


@router.message(Command("get_users_who_launched_the_bot"))
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.clear()  # Завершаем текущее состояние машины состояний
    try:
        if message.from_user.id not in [535185511, 301634256, 244948554, 496261899]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        orders = reading_from_database()
        workbook = create_excel_file_start(orders)  # Создание файла Excel
        filename = 'Данные_пользователей_запустивших_бота.xlsx'
        workbook.save(filename)  # Сохранение файла
        file = FSInputFile(filename)
        text = (
            "📊 <b>Данные пользователей, зарегистрированных в боте:</b>\n\n"
            "Файл содержит информацию о пользователях, которые запустили бота.\n\n"
            "⬇️ <b>Скачайте файл</b>, чтобы просмотреть данные.\n\n"
            "Для возврата в начальное меню нажмите на /admin_start."
        )
        await bot.send_document(message.from_user.id, document=file, caption=text,
                                parse_mode="HTML")  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(f"Произошла ошибка при попытке отправить файл: {e}")
        await bot.send_message(message.from_user.id,
                               "⚠️ Произошла ошибка при обработке вашего запроса. Пожалуйста, попробуйте позже.")


def register_greeting_admin_handler():
    """Регистрация обработчиков для бота"""
    dp.message.register(admin_start_handler)
    dp.message.register(admin_send_start)
    dp.message.register(get_users_get_photos_goods)


if __name__ == '__main__':
    reading_from_database()
    create_zip_archive()
    create_zip_archive_photos_goods()

